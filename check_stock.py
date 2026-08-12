#!/usr/bin/env python3
"""eBay在庫管理シートの在庫巡回スクリプト（GitHub Actions用）。

スプレッドシートの「在庫状況=在庫有り」行の仕入先URLを実際に確認し、
売り切れ・リンク切れ・判定不能があればLINEに通知する。

環境変数:
  SHEET_ID            スプレッドシートID（利益計算表ワークブック）
  SHEET_ID_SUB        サブ垢シートID
  MAIN_SOURCE         "rising" で利益計算表RisingJapanタブを本垢ソースとして使用
  LINE_CHANNEL_ID     LINE Messaging APIチャネルID
  LINE_CHANNEL_SECRET LINEチャネルシークレット
  LINE_USER_ID        通知先ユーザーID
  DRY_RUN             "1" ならLINE送信せずログ出力のみ
  DATA_DIR            state.json等のデータファイルを置くディレクトリ（既定: カレントディレクトリ）
"""
from __future__ import annotations

import base64
import csv
import html as html_mod
import io
import json
import os
import re
import signal
import socket
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from concurrent.futures import ThreadPoolExecutor
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone

from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import ec
from cryptography.hazmat.primitives.asymmetric.utils import decode_dss_signature

JST = timezone(timedelta(hours=9))
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")

# プロセス全体のソケットに強制タイムアウトをかける保険（1段目）。
socket.setdefaulttimeout(35)


class HardTimeout(Exception):
    pass


@contextmanager
def hard_timeout(seconds: int):
    """signal.alarmによる真の強制タイムアウト（2段目・本命）。
    2026-08-02、socket.setdefaulttimeout(35)だけでは不十分と判明——
    サーバーが極端に遅くデータを送り続ける("slowloris"的な)応答だと、
    urllibのtimeout引数は個々のread待ち時間しか区切らず、
    read()呼び出し全体が無期限に伸びてハングし続ける事象を実測。
    signal.alarmはブロッキング内容を問わずOSレベルで強制的に例外を発生させるため、
    この種のハングも確実に打ち切れる。
    【注意】signal.alarmはメインスレッドでしか使えない（他スレッドから呼ぶとValueError）。
    2026-08-02、ThreadPoolExecutor内のワーカースレッドから呼んで22件同時に
    ValueError誤判定を出す事故を実測。ワーカースレッドでは何もせず素通りする
    （その場合はsocket.setdefaulttimeout(35)の1段目のみで守られる）。"""
    if threading.current_thread() is not threading.main_thread():
        yield
        return
    def _on_alarm(signum, frame):
        raise HardTimeout(f"{seconds}秒のハードタイムアウト")
    old = signal.signal(signal.SIGALRM, _on_alarm)
    signal.alarm(seconds)
    try:
        yield
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old)

IN_STOCK = "IN_STOCK"
OUT_OF_STOCK = "OUT_OF_STOCK"
LINK_DEAD = "LINK_DEAD"
UNKNOWN = "UNKNOWN"


def log(msg: str) -> None:
    print(f"[{datetime.now(JST).strftime('%H:%M:%S')}] {msg}", flush=True)


# ---------------------------------------------------------------- sheet
def _fetch_csv(sheet_id: str, gid: str) -> list[list[str]]:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=60) as res:
        return list(csv.reader(io.StringIO(res.read().decode("utf-8"))))


def fetch_sheet(sheet_id: str) -> tuple[list[dict], dict]:
    """本垢の在庫管理表。(巡回対象の行リスト, 全行のキー→在庫状況マップ) を返す。
    マップのキーは「口座::URL」形式。"""
    rows, all_status = [], {}
    for r in _fetch_csv(sheet_id, "0"):
        if len(r) < 8:
            continue
        stock, url_, sku, name = r[4].strip(), r[3].strip(), r[7].strip(), r[1].strip()
        if not url_.startswith("http"):
            continue
        all_status[f"本垢::{url_}"] = {"sku": sku or "SKU不明", "stock": stock,
                                       "name": name.replace("\n", " ")}
        if stock == "在庫有り":
            rows.append({"sku": sku or "SKU不明", "name": name.replace("\n", " "),
                         "url": url_, "account": "本垢"})
    return rows, all_status


def fetch_sheet_rising(sheet_id: str) -> tuple[list[dict], dict]:
    """本垢の新ソース: 利益計算表ブックの「RisingJapan」タブ（2026-08-01から使用）。
    キーワードセル（E列）が赤=end list なので除外し、赤以外×URLありを巡回対象とする。
    色はCSVでは取れないためxlsxでダウンロードしてopenpyxlで読む。
    列: C=SKU, D=ItemID, E=キーワード, G=EC site URL, L=仕入価格。"""
    import openpyxl  # GitHub Actions側は workflow で pip install する
    import tempfile
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as res:
        raw = res.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(raw)
        path = f.name
    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        ws = wb["RisingJapan"]
        rows, all_status = [], {}
        for r in ws.iter_rows(min_row=8):
            sku = str(r[2].value or "").strip()
            url_ = str(r[6].value or "").strip()
            if not sku or not url_.startswith("http"):
                continue
            name = str(r[4].value or "").strip().replace("\n", " ")
            fill = r[4].fill
            rgb = str(fill.fgColor.rgb) if fill and fill.patternType and fill.fgColor else ""
            is_red = any(c in rgb for c in ("FF0000", "CC0000", "E06666", "EA4335"))
            # I列(9列目)= 在庫判定列。「在」「有在庫」=ユーザーが自宅に実物確保→巡回不要・上書き禁止
            i_val = str(r[8].value or "").strip()
            is_mine = i_val in ("在", "有在庫")
            all_status[f"本垢::{url_}"] = {
                "sku": sku, "name": name,
                "stock": "赤" if is_red else ("在" if is_mine else "有"),
            }
            if not is_red and not is_mine:
                # TODO: 8月に検証ログを月次ローテーションした後、r[11]の仕入価格で
                # 本垢も価格±5%監視を有効化する（現状はログ列数の互換性のため見送り）
                rows.append({"sku": sku, "name": name, "url": url_, "account": "本垢",
                             "sheet_row": r[0].row, "i_val": i_val})
        return rows, all_status
    finally:
        os.unlink(path)


def fetch_sheet_sub(sheet_id: str, gid: str) -> tuple[list[dict], dict]:
    """サブ垢（時計）の利益計算表Sakurarisingタブ。ヘッダーは7行目、
    列: C=SKU(wat…), E=商品名, G=EC site URL, I=有無(有/無/消/在), L=仕入価格。
    2026-08-07〜: 本垢と同じく「商品名セル(E列)が赤=出品取り下げ済み」を巡回除外の
    基準にする。加えて2026-08-10〜: 本垢の「在」「有在庫」と同じ意味で、I列が
    「在」「有在庫」の行（自宅に実物確保済み）も巡回除外にする（wat035で導入）。
    色はCSVでは取れないためxlsxでダウンロードしてopenpyxlで読む。"""
    import openpyxl
    import tempfile
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as res:
        raw = res.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(raw)
        path = f.name
    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        ws = wb["Sakurarising"]
        rows, all_status = [], {}
        for r in ws.iter_rows(min_row=8):
            sku = str(r[2].value or "").strip()
            url_ = str(r[6].value or "").strip()
            if not sku or not url_.startswith("http"):
                continue
            name = str(r[4].value or "").strip().replace("\n", " ")
            fill = r[4].fill
            rgb = str(fill.fgColor.rgb) if fill and fill.patternType and fill.fgColor else ""
            is_red = any(c in rgb for c in ("FF0000", "CC0000", "E06666", "EA4335"))
            i_val = str(r[8].value or "").strip()
            is_mine = i_val in ("在", "有在庫")
            all_status[f"サブ垢::{url_}"] = {
                "sku": sku, "name": name,
                "stock": "赤" if is_red else ("在" if is_mine else "有"),
            }
            if not is_red and not is_mine:
                price_text = str(r[11].value or "")
                m = re.search(r"[\d,]+", price_text)
                reg_price = int(m.group(0).replace(",", "")) if m else None
                rows.append({"sku": sku, "name": name, "url": url_,
                             "account": "サブ垢", "reg_price": reg_price})
        return rows, all_status
    finally:
        os.unlink(path)


# ---------------------------------------------------------------- mercari API
MERCARI_API = "https://api.mercari.jp/items/get"
MERCARI_ITEM_RE = re.compile(r"jp\.mercari\.com/item/(m\d+)")


def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode()


def _mercari_dpop() -> str:
    key = ec.generate_private_key(ec.SECP256R1())
    pub = key.public_key().public_numbers()
    jwk = {"crv": "P-256", "kty": "EC",
           "x": _b64url(pub.x.to_bytes(32, "big")),
           "y": _b64url(pub.y.to_bytes(32, "big"))}
    header = {"typ": "dpop+jwt", "alg": "ES256", "jwk": jwk}
    payload = {"iat": int(time.time()), "jti": str(uuid.uuid4()),
               "htu": MERCARI_API, "htm": "GET", "uuid": str(uuid.uuid4())}
    signing_input = _b64url(json.dumps(header).encode()) + "." + _b64url(json.dumps(payload).encode())
    der = key.sign(signing_input.encode(), ec.ECDSA(hashes.SHA256()))
    r, s = decode_dss_signature(der)
    return signing_input + "." + _b64url(r.to_bytes(32, "big") + s.to_bytes(32, "big"))


def _mercari_api_once(item_id: str) -> tuple[str, str, int | None]:
    req = urllib.request.Request(
        f"{MERCARI_API}?id={item_id}",
        headers={"DPoP": _mercari_dpop(), "X-Platform": "web",
                 "User-Agent": UA, "Accept": "application/json"})
    with hard_timeout(40), urllib.request.urlopen(req, timeout=30) as res:
        data = json.load(res).get("data", {})
    status = data.get("status", "")
    price = data.get("price") if isinstance(data.get("price"), int) else None
    if status == "on_sale":
        return IN_STOCK, "on_sale", price
    if status in ("sold_out", "trading"):
        return OUT_OF_STOCK, status, None
    if status in ("stop", "cancel", "admin_cancel", "deleted"):
        return LINK_DEAD, f"出品停止({status})", None
    return UNKNOWN, f"未知のstatus:{status}", None


# 403リトライの累計回数。上限を超えたら残りは次回に回す（実行時間の暴走防止）
MERCARI_RETRY_BUDGET = int(os.environ.get("MERCARI_RETRY_BUDGET", "5"))
_mercari_retries_used = 0


def check_mercari(item_id: str) -> tuple[str, str, int | None]:
    """403(レート制限)は待って1回リトライ。404はリンク切れ。3要素目は現在価格。"""
    global _mercari_retries_used
    for attempt in (1, 2):
        try:
            return _mercari_api_once(item_id)
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return LINK_DEAD, "メルカリAPI 404 商品なし", None
            if e.code == 403 and attempt == 1 and _mercari_retries_used < MERCARI_RETRY_BUDGET:
                _mercari_retries_used += 1
                log(f"  メルカリAPI 403 → 90秒クールダウン ({item_id}, "
                    f"リトライ{_mercari_retries_used}/{MERCARI_RETRY_BUDGET})")
                time.sleep(90)
                continue
            return (UNKNOWN, "メルカリAPIレート制限", None) if e.code == 403 \
                else (UNKNOWN, f"メルカリAPI HTTP {e.code}", None)
        except Exception as e:  # noqa: BLE001
            return UNKNOWN, f"メルカリAPI {type(e).__name__}", None
    return UNKNOWN, "メルカリAPIレート制限", None


# ---------------------------------------------------------------- HTTP fetch
def fetch_html(url: str, retry: bool = True) -> tuple[int, str]:
    """(HTTPステータス, HTML) を返す。429/5xx/タイムアウトは1回リトライ。"""
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ja,en-US;q=0.7",
    })
    try:
        with hard_timeout(40), urllib.request.urlopen(req, timeout=30) as res:
            raw = res.read()
            charset = res.headers.get_content_charset()
            if not charset:
                m = re.search(rb'charset=["\']?([\w-]+)', raw[:2000])
                charset = m.group(1).decode("ascii", "replace") if m else "utf-8"
            try:
                return res.status, raw.decode(charset, errors="replace")
            except LookupError:
                return res.status, raw.decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        if e.code in (429, 500, 502, 503) and retry:
            time.sleep(60)
            return fetch_html(url, retry=False)
        return e.code, ""
    except Exception as e:  # noqa: BLE001
        if retry:
            time.sleep(15)
            return fetch_html(url, retry=False)
        return 0, f"__FETCH_ERROR__{type(e).__name__}"


# ---------------------------------------------------------------- site rules
AVAILABILITY_RE = re.compile(r'https?://schema\.org/(InStock|OutOfStock|SoldOut|Discontinued)')


def _jsonld_product_availability(html: str) -> str | None:
    """メインのProduct JSON-LDからavailabilityを取り出す（ghibli等）。"""
    for m in re.findall(r'<script type="application/ld\+json">(.*?)</script>', html, re.S):
        try:
            d = json.loads(html_mod.unescape(m).strip())
        except Exception:  # noqa: BLE001
            continue
        items = d if isinstance(d, list) else [d]
        for item in items:
            if isinstance(item, dict) and item.get("@type") == "Product":
                offers = item.get("offers") or {}
                if isinstance(offers, list):
                    offers = offers[0] if offers else {}
                av = offers.get("availability", "") if isinstance(offers, dict) else str(offers)
                m2 = AVAILABILITY_RE.search(str(av))
                if m2:
                    return m2.group(1)
    return None


def _avail_name(text: str) -> str | None:
    m = AVAILABILITY_RE.search(text)
    return m.group(1) if m else None


def _from_availability(name: str | None, detail: str) -> tuple[str, str] | None:
    if name is None:
        return None
    if name == "InStock":
        return IN_STOCK, f"{detail}=InStock"
    return OUT_OF_STOCK, f"{detail}={name}"


def rule_disney(html: str) -> tuple[str, str] | None:
    m = re.search(r'availability"[^>]*>\s*(https?://schema\.org/\w+)', html)
    if m:
        return _from_availability(_avail_name(m.group(1)), "availability")
    return None


def rule_jsonld(html: str) -> tuple[str, str] | None:
    return _from_availability(_jsonld_product_availability(html), "JSON-LD")


def rule_schema_or_flag(html: str) -> tuple[str, str] | None:
    """yahooショッピング・USJ等: ページ内JSONのフラグで判定。"""
    if '"stock":{"isAvailable":true' in html or '"isAvailable":true' in html:
        return IN_STOCK, "isAvailable=true"
    if '"stock":{"isAvailable":false' in html or '"isAvailable":false' in html:
        return OUT_OF_STOCK, "isAvailable=false"
    m = re.search(r'"availability":"(https?://schema\.org/\w+)"', html)
    if m:
        return _from_availability(_avail_name(m.group(1)), "availability")
    return None


def rule_amazon(html: str) -> tuple[str, str] | None:
    low = html.lower()
    if "api-services-support@amazon.com" in low or "robot check" in low or "検証が必要です" in html:
        return UNKNOWN, "Amazon bot対策ページ"
    # availability div内の「最初の表示文」だけで判定する。
    # ページ全体やdiv周辺1200文字を見ると、おすすめ商品カルーセルの
    # 「一時的に在庫切れ」等を誤って拾う（2026-07-13 fis010誤検知の原因）
    m = re.search(r'id="availability"[^>]*>.*?<span[^>]*>\s*([^<]{3,120})', html, re.S)
    if m:
        text = m.group(1).strip()
        if any(k in text for k in ("在庫あり", "以内に発送", "残り", "まもなく発売")):
            return IN_STOCK, f"availability欄「{text[:20]}」"
        if any(k in text for k in ("一時的に在庫切れ", "お取り扱いできません", "在庫切れ")):
            return OUT_OF_STOCK, f"availability欄「{text[:20]}」"
    if 'id="add-to-cart-button"' in html:
        return IN_STOCK, "カートボタンあり"
    if "この商品は現在お取り扱いできません" in html:
        return OUT_OF_STOCK, "取り扱い終了表示"
    return None


def rule_paypay(html: str) -> tuple[str, str] | None:
    if '"status":"OPEN"' in html:
        return IN_STOCK, "status=OPEN"
    for s in ("SOLD_OUT", "SOLD", "CLOSED"):
        if f'"status":"{s}"' in html:
            return OUT_OF_STOCK, f"status={s}"
    return None


def rule_fril(html: str) -> tuple[str, str] | None:
    if "すぐに購入可" in html or "購入画面に進む" in html:
        return IN_STOCK, "購入ボタンあり"
    if "売り切れました" in html or "SOLD OUT" in html:
        return OUT_OF_STOCK, "売り切れ表示"
    return None


def rule_pokemon(html: str) -> tuple[str, str] | None:
    m = re.search(r'<button[^>]*add-to-cart[^>]*>', html)
    if m:
        if "disabled" in m.group(0):
            return OUT_OF_STOCK, "カートボタン無効(品切れ)"
        return IN_STOCK, "カートボタン有効"
    return None


def rule_rakuten(html: str) -> tuple[str, str] | None:
    if "'soldout':[0]" in html:
        return IN_STOCK, "soldout=0"
    if "'soldout':[1]" in html:
        return OUT_OF_STOCK, "soldout=1"
    # 新形式ストアフロント（ビックカメラ楽天市場店など）は埋め込みJSONのフラグで判定
    if '"sold_out_flag":false' in html:
        m = re.search(r'"inventory":(\d+)', html)
        return IN_STOCK, f"sold_out_flag=false" + (f"(在庫{m.group(1)})" if m else "")
    if '"sold_out_flag":true' in html:
        return OUT_OF_STOCK, "sold_out_flag=true"
    return None


def rule_donguri(html: str) -> tuple[str, str] | None:
    m = re.search(r"HERE'>(\d+)</span>\s*点", html)
    if m:
        qty = int(m.group(1))
        return (IN_STOCK, f"在庫数{qty}点") if qty > 0 else (OUT_OF_STOCK, "在庫数0点")
    return None


def rule_auctions(html: str) -> tuple[str, str] | None:
    if "このオークションは終了しています" in html or "落札されました" in html:
        return OUT_OF_STOCK, "オークション終了"
    if "入札する" in html or "残り時間" in html:
        return IN_STOCK, "オークション開催中"
    return None


# サイト別: 適用するルールのリスト（先勝ち）
SITE_HANDLERS: dict[str, list] = {
    "www.amazon.co.jp": [rule_amazon],
    "store.disney.co.jp": [rule_disney, rule_jsonld],
    "www.ghibli-museum-shop.jp": [rule_jsonld],
    "store.shopping.yahoo.co.jp": [rule_schema_or_flag, rule_jsonld],
    "paypayfleamarket.yahoo.co.jp": [rule_paypay],
    "item.fril.jp": [rule_fril],
    "www.pokemoncenter-online.com": [rule_pokemon, rule_jsonld],
    "auctions.yahoo.co.jp": [rule_auctions],
    "www.onlinestore.usj.co.jp": [rule_schema_or_flag, rule_jsonld],
    "item.rakuten.co.jp": [rule_rakuten, rule_jsonld, rule_schema_or_flag],
    "biccamera.rakuten.co.jp": [rule_rakuten, rule_jsonld, rule_schema_or_flag],
    "www.donguri-sora.com": [rule_donguri],
}

# 汎用マーカー（専用ルールで判定できなかった場合）
GENERIC_IN = ["カートに入れる", "カートに追加", "かごに追加", "買い物かごに入れる", "ご購入手続きへ", "カートに入れ"]
GENERIC_OUT = ["SOLD OUT", "売り切れました", "在庫がありません", "販売休止中", "完売しました", "販売を終了しました"]

# 恒常的に自動判定できないサイト（要個別対応）
MANUAL_SITES = {
    "www.suruga-ya.jp": "駿河屋はbot対策(403)のため自動判定不可",
}

# このIP環境からはチェックしないサイト（別環境の巡回に任せる）。カンマ区切り。
SKIP_HOSTS = {h.strip() for h in os.environ.get("SKIP_HOSTS", "").split(",") if h.strip()}
SKIPPED = "SKIPPED"


def check_generic(url: str) -> tuple[str, str]:
    host = urllib.parse.urlparse(url).netloc
    if host in SKIP_HOSTS:
        return SKIPPED, "別環境で巡回"
    if host in MANUAL_SITES:
        return UNKNOWN, MANUAL_SITES[host]
    if "jp.mercari.com/shops/" in url:
        return SKIPPED, "別環境で巡回"  # メルカリShopsはJSレンダリングのためローカル巡回が担当

    status, html = fetch_html(url)
    if status in (404, 410):
        # PayPayフリマ・ヤフオクはデータセンターIPに偽の404を返すことがある
        # （2026-07-14 hob885、2026-08-07 hob835: 実際は販売中/開催中だった）→断定せず要確認扱い
        if host in ("paypayfleamarket.yahoo.co.jp", "auctions.yahoo.co.jp"):
            return UNKNOWN, "HTTP 404（bot対策の可能性あり・要手動確認）"
        return LINK_DEAD, f"HTTP {status}"
    if html.startswith("__FETCH_ERROR__"):
        return UNKNOWN, f"接続失敗({html.removeprefix('__FETCH_ERROR__')})"
    if status != 200:
        return UNKNOWN, f"HTTP {status}"

    for handler in SITE_HANDLERS.get(host, []):
        result = handler(html)
        if result:
            return result

    # 汎用フォールバック: JSON-LD → マーカー
    result = rule_jsonld(html)
    if result:
        return result
    hit_in = [m for m in GENERIC_IN if m in html]
    hit_out = [m for m in GENERIC_OUT if m in html]
    if hit_out and not hit_in:
        return OUT_OF_STOCK, f"「{hit_out[0]}」検出"
    if hit_in and not hit_out:
        return IN_STOCK, f"「{hit_in[0]}」検出"
    if hit_in and hit_out:
        return UNKNOWN, f"在庫/売切の両表示が混在({hit_in[0]}/{hit_out[0]})"
    return UNKNOWN, "在庫表示が見つからない"


# ---------------------------------------------------------------- dispatch
PRICE_ALERT_PCT = float(os.environ.get("PRICE_ALERT_PCT", "5")) / 100


def check_one(row: dict) -> dict:
    url = row["url"]
    m = MERCARI_ITEM_RE.search(url)
    price = None
    if m:
        cat, reason, price = check_mercari(m.group(1))
    else:
        cat, reason = check_generic(url)
    out = {**row, "category": cat, "reason": reason,
           "site": urllib.parse.urlparse(url).netloc}
    # 登録価格(仕入価格)がある行は現在価格との乖離を監視（現状はメルカリのみ価格取得可）
    reg = row.get("reg_price")
    if price and reg:
        dev = (price - reg) / reg
        if abs(dev) >= PRICE_ALERT_PCT:
            out["price_alert"] = f"¥{reg:,}→¥{price:,}（{dev:+.1%}）"
    return out


# ---------------------------------------------------------------- LINE
def send_line(text: str) -> None:
    if os.environ.get("DRY_RUN") == "1":
        log("DRY_RUN=1 のためLINE送信スキップ。送信予定の内容:")
        print(text, flush=True)
        return
    data = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": os.environ["LINE_CHANNEL_ID"],
        "client_secret": os.environ["LINE_CHANNEL_SECRET"],
    }).encode("ascii")
    req = urllib.request.Request(
        "https://api.line.me/oauth2/v3/token", data=data,
        headers={"Content-Type": "application/x-www-form-urlencoded"}, method="POST")
    with urllib.request.urlopen(req, timeout=30) as res:
        token = json.loads(res.read())["access_token"]
    body = json.dumps({
        "to": os.environ["LINE_USER_ID"],
        "messages": [{"type": "text", "text": text[:4900]}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.line.me/v2/bot/message/push", data=body,
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {token}"}, method="POST")
    with urllib.request.urlopen(req, timeout=30) as res:
        log(f"LINE送信 status={res.status}")


def build_account_digest(account: str, results: list[dict], all_items: list[dict]) -> str | None:
    """1口座分の日次レポートを作る。対象商品が0件ならNone（送信しない）。
    all_itemsは赤(除外)以外の全商品(sku/name)。メルカリのローテーション等で
    直近24時間にログが無い商品もあるため、実際にチェックできた件数(len(acc))とは
    別に総数を明記し（2026-08-10）、さらにチェック漏れのSKUも列挙する
    （2026-08-10、「対象と実チェック数の差が何か分からない」という指摘に対応）。"""
    total = len(all_items)
    acc = [r for r in results if r["account"] == account]
    if not acc and total == 0:
        return None
    sold = [r for r in acc if r["category"] == OUT_OF_STOCK]
    dead = [r for r in acc if r["category"] == LINK_DEAD]
    unknown = [r for r in acc if r["category"] == UNKNOWN
               and r["reason"] != "メルカリAPIレート制限"]
    rate_limited = [r for r in acc if r["reason"] == "メルカリAPIレート制限"]
    price_alerts = [r for r in acc if r["price_alert"] and r["category"] == IN_STOCK]
    ok = sum(1 for r in acc if r["category"] == IN_STOCK)

    lines = [f"【在庫巡回 日次レポート（{account}） {datetime.now(JST).strftime('%-m/%-d')}】",
             f"対象{total}件中{len(acc)}件チェック（在庫あり{ok}件）"]
    if not (sold or dead or unknown or price_alerts):
        lines.append("問題は見つかりませんでした。")

    def section(title: str, items: list[dict], with_reason: bool) -> None:
        if not items:
            return
        # 同一SKUの重複行（シートに同じ商品が複数行ある場合）は1つにまとめる
        seen: set[str] = set()
        items = [r for r in items if not (r["sku"] in seen or seen.add(r["sku"]))]
        lines.append(f"■{title}")
        # 同一サイト×同一原因が6件以上ならまとめる（レート制限時のスパム防止）。SKUは必ず列挙する
        if with_reason:
            from collections import Counter
            keys = Counter((r["site"], r["reason"]) for r in items)
            bulk = {k for k, n in keys.items() if n >= 6}
            for site, reason in sorted(bulk):
                skus = [r["sku"] for r in items if (r["site"], r["reason"]) == (site, reason)]
                lines.append(f"・{reason} ×{len(skus)}件: {', '.join(skus)}")
            items = [r for r in items if (r["site"], r["reason"]) not in bulk]
        for r in items:
            entry = f"・{r['sku']} {r['name'][:15]}"
            if with_reason:
                entry += f"（{r['site']}: {r['reason']}）"
            lines.append(entry)
            if with_reason:
                # 手動確認しやすいよう商品ページのURLを併記（Amazonは短縮形に）
                m = re.search(r"amazon\.co\.jp.*?/dp/([A-Z0-9]{10})", r["url"])
                lines.append(f"  https://www.amazon.co.jp/dp/{m.group(1)}" if m
                             else f"  {r['url'][:250]}")

    section("売り切れ（シート更新をお願いします）", sold, False)
    section("リンク切れ・出品削除（シート更新をお願いします）", dead, True)
    if price_alerts:
        lines.append(f"■価格変動±{PRICE_ALERT_PCT:.0%}以上（仕入価格の見直しを）")
        for r in price_alerts:
            lines.append(f"・{r['sku']} {r['name'][:15]} {r['price_alert']}")
            lines.append(f"  {r['url'][:250]}")
    section("判定不能（お手すきに目視確認を）", unknown, True)
    if rate_limited:
        lines.append(f"※メルカリ混雑で{len(rate_limited)}件は未チェック"
                     f"（自動で再確認します・対応不要）: "
                     + ", ".join(sorted(r["sku"] for r in rate_limited)))
    # 対象件数はあるがログが無い＝直近24時間チェックできていない商品（メルカリ回転制等）。
    # 「対象◯件中◯件チェック」という表記だけでは差分の中身が分からないとの指摘に対応し、
    # SKUを列挙する（在庫なし確定ではなく、あくまで「まだ確認できていない」の意）。
    checked_skus = {r["sku"] for r in acc} | {r["sku"] for r in rate_limited}
    unchecked = sorted({it["sku"] for it in all_items if it["sku"] not in checked_skus})
    if unchecked:
        lines.append(f"※直近24時間チェックできていない{len(unchecked)}件"
                     f"（在庫なし確定ではありません・次回巡回で確認します）: "
                     + ", ".join(unchecked))
    lines.append("―――")
    lines.append("間違いがあればClaudeのチャットにSKUと実際の状態を送ってください（例:「hob885 在庫あり」）")
    return "\n".join(lines)


def build_daily_digest(all_status: dict) -> list[str]:
    """直近24時間の検証ログ（GitHub分＋ローカル分、本垢＋サブ垢）を集約し、
    口座ごとに独立したLINEメッセージのリストを返す（該当商品が0件の口座は含めない）。
    各商品はその24時間内の最新判定を採用する。"""
    cutoff = datetime.now(JST) - timedelta(hours=24)
    latest: dict = {}
    sources = [(VERIFY_LOG, "本垢"),
               (os.path.join(BASE_DIR, "verify_log_local.csv"), "本垢"),
               (VERIFY_LOG_SUB, "サブ垢"),
               (os.path.join(BASE_DIR, "verify_log_local_sub.csv"), "サブ垢")]
    for path, account in sources:
        if not os.path.exists(path):
            continue
        with open(path, newline="") as f:
            for row in csv.DictReader(f):
                try:
                    t = datetime.strptime(row["datetime_jst"], "%Y-%m-%d %H:%M").replace(tzinfo=JST)
                except (ValueError, TypeError):
                    continue
                if t < cutoff:
                    continue
                key = (account, row["url"])
                if key not in latest or latest[key][0] <= t:
                    latest[key] = (t, row, account)

    def to_item(row: dict, account: str) -> dict | None:
        url = row["url"]
        info = all_status.get(f"{account}::{url}")
        # シートからURLが消えた／ユーザーが既に売切等へ更新した行は報告しない
        # （URL差し替え後に旧URLの判定が24時間残って誤報になるのを防ぐ。2026-07-19 hob465）
        if info is None or info.get("stock") not in ("在庫有り", "有"):
            return None
        return {"sku": row["sku"], "url": url, "account": account,
                "name": info.get("name", ""),
                "site": urllib.parse.urlparse(url).netloc,
                "reason": row["reason"], "category": row["category"],
                "price_alert": (row.get("price_alert") or "").strip()}

    results = [it for _, row, account in latest.values()
               if (it := to_item(row, account)) is not None]

    all_items: dict[str, list[dict]] = {"本垢": [], "サブ垢": []}
    for key, info in all_status.items():
        account = key.split("::", 1)[0]
        if account in all_items and info.get("stock") in ("在庫有り", "有"):
            all_items[account].append({"sku": info["sku"], "name": info.get("name", "")})

    return [msg for account in ("本垢", "サブ垢")
            if (msg := build_account_digest(account, results, all_items[account])) is not None]


# ---------------------------------------------------------------- main
# DATA_DIRが指定されていればそちら（GitHub Actions側でprivateなdataリポジトリを
# クローンしたディレクトリ）を、無ければスクリプト自身のディレクトリ（ローカル実行時）を使う。
BASE_DIR = os.environ.get("DATA_DIR") or os.path.dirname(os.path.abspath(__file__))
STATE_PATH = os.path.join(BASE_DIR, "state.json")
# メルカリAPIは1回の実行で約90件以上叩くとレート制限になるため、
# 1実行あたりの件数を絞り、続きの位置をstate.jsonに記録してローテーションする
MERCARI_PER_RUN = int(os.environ.get("MERCARI_PER_RUN", "80"))


def load_state() -> dict:
    try:
        with open(STATE_PATH) as f:
            return json.load(f)
    except Exception:  # noqa: BLE001
        return {}


def save_state(state: dict) -> None:
    with open(STATE_PATH, "w") as f:
        json.dump(state, f, ensure_ascii=False, indent=1)


# ---------------------------------------------------------------- 検証ログ（在庫管理ツール解約判断用）
VERIFY_LOG = os.path.join(BASE_DIR, "verify_log.csv")
VERIFY_LOG_SUB = os.path.join(BASE_DIR, "verify_log_sub.csv")
TRANSITIONS_LOG = os.path.join(BASE_DIR, "transitions.csv")
SNAPSHOT_PATH = os.path.join(BASE_DIR, "sheet_snapshot.json")


def append_verify_log(results: list[dict]) -> None:
    """全チェック結果を追記（月末にAIの検出精度を集計するための生データ）。
    本垢はverify_log.csv（5列）、サブ垢はverify_log_sub.csv（価格変動列つき6列）。"""
    ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M")
    # SKIPPED(別環境巡回待ち)は情報を持たないため記録しない
    results = [r for r in results if r["category"] != SKIPPED]
    main_rows = [r for r in results if r.get("account", "本垢") == "本垢"]
    sub_rows = [r for r in results if r.get("account") == "サブ垢"]
    if main_rows:
        new_file = not os.path.exists(VERIFY_LOG)
        with open(VERIFY_LOG, "a", newline="") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(["datetime_jst", "sku", "url", "category", "reason"])
            for r in main_rows:
                w.writerow([ts, r["sku"], r["url"], r["category"], r["reason"]])
    if sub_rows:
        new_file = not os.path.exists(VERIFY_LOG_SUB)
        with open(VERIFY_LOG_SUB, "a", newline="") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(["datetime_jst", "sku", "url", "category", "reason", "price_alert"])
            for r in sub_rows:
                w.writerow([ts, r["sku"], r["url"], r["category"], r["reason"],
                            r.get("price_alert", "")])


def record_tool_transitions(all_status: dict) -> None:
    """在庫管理ツールがシートを「在庫有り→売切等」に変えた行を検出し、
    AIが直近3日以内に先回りして検出できていたかを記録する。"""
    try:
        with open(SNAPSHOT_PATH) as f:
            prev = json.load(f)
    except Exception:  # noqa: BLE001
        prev = {}

    if prev:
        # 直近3日のAI検出履歴（OUT_OF_STOCK/LINK_DEAD）を読み込む
        detected_urls: set[str] = set()
        cutoff = datetime.now(JST) - timedelta(days=3)
        for path in (VERIFY_LOG, VERIFY_LOG_SUB):
            if not os.path.exists(path):
                continue
            with open(path, newline="") as f:
                for row in csv.DictReader(f):
                    if row["category"] in (OUT_OF_STOCK, LINK_DEAD):
                        try:
                            t = datetime.strptime(row["datetime_jst"], "%Y-%m-%d %H:%M").replace(tzinfo=JST)
                        except ValueError:
                            continue
                        if t >= cutoff:
                            detected_urls.add(row["url"])
        changes = []
        ACTIVE = ("在庫有り", "有")  # 本垢は「在庫有り」、サブ垢は「有」
        for key, info in all_status.items():
            url = key.split("::", 1)[-1]
            p = prev.get(key) or prev.get(url)  # 旧形式(URLのみキー)の互換
            if p and p.get("stock") in ACTIVE and info["stock"] not in ACTIVE:
                changes.append([datetime.now(JST).strftime("%Y-%m-%d %H:%M"),
                                info["sku"], url, f"{p.get('stock')}→{info['stock']}",
                                "先に検出済み" if url in detected_urls else "AI未検出"])
        if changes:
            new_file = not os.path.exists(TRANSITIONS_LOG)
            with open(TRANSITIONS_LOG, "a", newline="") as f:
                w = csv.writer(f)
                if new_file:
                    w.writerow(["datetime_jst", "sku", "url", "tool_change", "ai_result"])
                w.writerows(changes)
            for c in changes:
                log(f"  ツール側変更: {c[1]} {c[3]} → {c[4]}")

    with open(SNAPSHOT_PATH, "w") as f:
        json.dump(all_status, f, ensure_ascii=False)


def main() -> int:
    sheet_id = os.environ["SHEET_ID"]
    # 本垢のソース: legacy=在庫管理表(〜2026-07-31) / rising=利益計算表RisingJapanタブ(2026-08-01〜)
    if os.environ.get("MAIN_SOURCE", "legacy") == "rising":
        rows, all_status = fetch_sheet_rising(os.environ["SHEET_ID_SUB"] or sheet_id)
    else:
        rows, all_status = fetch_sheet(sheet_id)
    sub_id = os.environ.get("SHEET_ID_SUB", "").strip()
    if sub_id:
        sub_rows, sub_status = fetch_sheet_sub(sub_id, os.environ.get("SHEET_SUB_GID", "1368802580"))
        rows += sub_rows
        all_status.update(sub_status)
        log(f"巡回対象: 本垢{len(rows) - len(sub_rows)}件＋サブ垢{len(sub_rows)}件")
    else:
        log(f"巡回対象: {len(rows)}件")
    record_tool_transitions(all_status)

    mercari = [r for r in rows if MERCARI_ITEM_RE.search(r["url"])]
    others = [r for r in rows if not MERCARI_ITEM_RE.search(r["url"])]

    # メルカリはローテーション: 前回の続きからMERCARI_PER_RUN件だけチェック。
    # 前回レート制限でチェックできなかった商品(pending)は最優先で先頭に入れる
    state = load_state()
    def item_id(r: dict) -> str:
        return MERCARI_ITEM_RE.search(r["url"]).group(1)
    pending_ids = set(state.get("pending_mercari", []))
    if len(mercari) > MERCARI_PER_RUN:
        pending_rows = [r for r in mercari if item_id(r) in pending_ids]
        rest = [r for r in mercari if item_id(r) not in pending_ids]
        offset = state.get("mercari_offset", 0) % max(len(rest), 1)
        rotated = rest[offset:] + rest[:offset]
        quota = max(MERCARI_PER_RUN - len(pending_rows), 0)
        mercari = pending_rows + rotated[:quota]
        state["mercari_offset"] = (offset + quota) % max(len(rest), 1)
        log(f"メルカリ: 優先再チェック{len(pending_rows)}件＋通常{quota}件"
            f"（オフセット{offset}、レート制限対策のローテーション）")

    results: list[dict] = []

    # その他サイト: ホストごとに直列（同一サイトへの連続アクセスを1.2秒空ける）、ホスト間は並列
    by_host: dict[str, list[dict]] = {}
    for r in others:
        by_host.setdefault(urllib.parse.urlparse(r["url"]).netloc, []).append(r)

    HOST_DELAY = {"paypayfleamarket.yahoo.co.jp": 4.0}

    def run_host(host: str, host_rows: list[dict]) -> list[dict]:
        delay = HOST_DELAY.get(host, 1.2)
        out = []
        for r in host_rows:
            out.append(check_one(r))
            time.sleep(delay)
        return out

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = [pool.submit(run_host, h, hr) for h, hr in by_host.items()]
        # メルカリAPIはメインスレッドで直列（2秒間隔。件数はローテーションで80件以下に制限済み）
        for i, r in enumerate(mercari, 1):
            results.append(check_one(r))
            time.sleep(3.0)
            if i % 40 == 0:
                log(f"  メルカリ {i}/{len(mercari)} 完了")
        # 403のまま残った商品の再分類: 自分より後にチェックした商品が成功しているなら
        # その時点でAPIは生きていた＝レート制限ではなく出品削除
        # （メルカリAPIは削除・公開停止済み商品に403を返す。2026-07-15判明）
        last_ok = -1
        for i, r in enumerate(results):
            if r["reason"] != "メルカリAPIレート制限":
                last_ok = i
        for i, r in enumerate(results):
            if r["reason"] == "メルカリAPIレート制限" and i < last_ok:
                r["category"] = LINK_DEAD
                r["reason"] = "メルカリAPI 403（出品削除の可能性・要確認）"

        for f in futures:
            results.extend(f.result())

    # 今回レート制限で未チェックだった商品を記録 → 次回の実行が最優先で再チェック
    state["pending_mercari"] = [item_id(r) for r in results
                                if r["reason"] == "メルカリAPIレート制限"
                                and MERCARI_ITEM_RE.search(r["url"])]
    save_state(state)

    counts = {c: sum(1 for r in results if r["category"] == c)
              for c in (IN_STOCK, OUT_OF_STOCK, LINK_DEAD, UNKNOWN, SKIPPED)}
    log(f"結果: {counts}")
    for r in results:
        if r["category"] not in (IN_STOCK, SKIPPED):
            log(f"  {r['category']:12} {r['sku']:8} {r['site']} {r['reason']}")

    append_verify_log(results)

    # LINE通知は1日1回、本垢・サブ垢を別メッセージで送る。
    # 重複防止はlast_digest_date（日付）のみで行う。
    # 2026-08-02: 以前は「3時以降のみ」という時刻制限もあったが、
    # 深夜の復旧作業時に手動実行しても送信されない事故の原因になったため撤廃。
    today = datetime.now(JST).strftime("%Y-%m-%d")
    if state.get("last_digest_date") != today:
        msgs = build_daily_digest(all_status)
        if not msgs:
            log("日次レポート: 対象0件のため送信なし")
        for msg in msgs:
            send_line(msg)
            time.sleep(3)  # LINE側で連続送信が1通に見えないよう間隔を空ける
        if os.environ.get("DRY_RUN") != "1":
            state["last_digest_date"] = today
            save_state(state)
    else:
        log("日次レポートは送信済みのため通知なし（結果はログに記録済み）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
